"""
scripts/loaders/file_loader.py

File discovery, SHA-256 hashing, Excel reading, and voter file reading.
Raw data is NEVER modified here; outputs are always new objects.
"""

import hashlib
import os
import zipfile
from pathlib import Path
from typing import Iterator

import openpyxl


def sha256_file(path: str | Path) -> str:
    """Return hex SHA-256 of a file, or 'ERROR' if unreadable."""
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception as e:
        return f"ERROR:{e}"


def file_size(path: str | Path) -> int:
    """Return file size in bytes, or -1 if unavailable."""
    try:
        return os.path.getsize(path)
    except Exception:
        return -1


def discover_files(directory: str | Path, extensions: list[str]) -> list[Path]:
    """
    Recursively discover files with given extensions in a directory.
    Returns sorted list of absolute Paths.
    """
    directory = Path(directory)
    found = []
    if not directory.is_dir():
        return found
    for ext in extensions:
        found.extend(directory.rglob(f"*{ext}"))
    return sorted(set(found))


def load_excel_workbook(path: str | Path) -> openpyxl.Workbook:
    """
    Load an xlsx/xls workbook read-only. Returns openpyxl Workbook.
    Raises FileNotFoundError or ValueError on failure.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError(f"Unsupported Excel format: {path.suffix}")
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except zipfile.BadZipFile as e:
        with open(path, "r", errors="ignore") as f:
            header = f.read(100)
            if "<?xml" in header or "<?mso-application" in header:
                raise ValueError(
                    f"File '{path.name}' is an XML Spreadsheet masquerading as XLSX. "
                    "Open it in Excel and 'Save As' a true .xlsx workbook before running."
                ) from e
        raise ValueError(f"File '{path.name}' is not a valid .xlsx file or is corrupted.") from e


def iter_excel_sheets(workbook: openpyxl.Workbook) -> Iterator[tuple[str, list[list]]]:
    """
    Yield (sheet_name, rows) for each sheet in workbook.
    rows is a list of lists (raw cell values).

    Uses pandas for reading to handle variable-column workbooks correctly.
    openpyxl read_only mode can return truncated rows when max_column
    metadata is missing or wrong (common in files exported from CA election
    reporting systems). Falls back to openpyxl if pandas fails.
    """
    # We need the file path to use pandas; extract it from the workbook if possible
    wb_path = getattr(workbook, "_archive", None)
    if wb_path is None:
        # Try to get it via the fp attribute used by openpyxl internals
        wb_path = getattr(workbook, "fp", None)
    if wb_path is not None:
        wb_path = getattr(wb_path, "name", None)

    if wb_path:
        try:
            import pandas as pd
            import warnings
            xl = pd.ExcelFile(wb_path)
            for name in workbook.sheetnames:
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        df = pd.read_excel(xl, sheet_name=name, header=None, dtype=object)
                    # Convert to list-of-lists, replacing pd.NA/NaN with None
                    rows = []
                    for _, row in df.iterrows():
                        rows.append([None if (v is None or (isinstance(v, float) and __import__('math').isnan(v))) else v
                                      for v in row])
                    yield name, rows
                    continue
                except Exception:
                    pass
            return
        except Exception:
            pass

    # Fallback: openpyxl row iteration (may truncate rows in some workbooks)
    for name in workbook.sheetnames:
        ws = workbook[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        yield name, rows



def load_voter_file(path: str | Path) -> list[dict]:
    """
    Load a voter file (.csv, .tsv, .txt, .zip containing one of those).
    Returns list of dicts (header row → keys).
    NOTE: Voter files are optional; caller handles missing gracefully.
    """
    import csv

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Voter file not found: {path}")

    def _read_stream(stream, delimiter=",") -> list[dict]:
        reader = csv.DictReader(stream, delimiter=delimiter)
        return [dict(row) for row in reader]

    suffix = path.suffix.lower()
    if suffix == ".zip":
        with zipfile.ZipFile(path) as zf:
            for name in zf.namelist():
                if name.lower().endswith((".csv", ".tsv", ".txt")):
                    delim = "\t" if name.lower().endswith(".tsv") else ","
                    with zf.open(name) as raw:
                        import io
                        text = io.TextIOWrapper(raw, encoding="utf-8-sig")
                        return _read_stream(text, delimiter=delim)
        raise ValueError("No CSV/TSV/TXT found inside ZIP")

    delim = "\t" if suffix in (".tsv", ".txt") else ","
    with open(path, encoding="utf-8-sig", newline="") as f:
        return _read_stream(f, delimiter=delim)
