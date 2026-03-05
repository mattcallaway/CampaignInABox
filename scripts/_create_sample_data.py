"""
Create synthetic detail.xlsx fixture for SAMPLE_COUNTY/MEASURE_A.
Two sheets:
  Sheet 1: Measure A (ballot measure, YES/NO, 5 precincts, 2 vote methods)
  Sheet 2: DA Race (candidate race, 3 candidates, same 5 precincts)
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent.parent  # Campaign In A Box/
out_dir = BASE_DIR / "votes" / "2024" / "CA" / "SAMPLE_COUNTY" / "MEASURE_A"
out_dir.mkdir(parents=True, exist_ok=True)
out_path = out_dir / "detail.xlsx"

wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────────────────────────────
# Sheet 1: Measure A — Ballot Measure
# ──────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Measure A"

headers1 = [
    "Precinct_ID", "Precinct Name", "Registered",
    "Mail YES", "Mail NO",
    "In Person YES", "In Person NO",
    "YES", "NO", "Ballots Cast",
]

# 5 synthetic precincts
precincts = [
    ("0001", "Oakdale North",    4200, 812, 310,  430, 165, 1242, 475, 1717),
    ("0002", "Riverside Park",   3875, 725, 385,  312, 198,  1037, 583, 1620),
    ("0003", "Downtown Central", 5120, 920, 440,  615, 285,  1535, 725, 2260),
    ("0004", "Hillcrest East",   2900, 510, 280,  280, 160,   790, 440, 1230),
    ("0005", "Westview Valley",  6300,1150, 520,  720, 310,  1870, 830, 2700),
]

# Style helper
header_font  = Font(bold=True, color="FFFFFF")
header_fill  = PatternFill("solid", fgColor="2D4A8A")
center_align = Alignment(horizontal="center")

ws1.append(headers1)
for cell in ws1[1]:
    cell.font  = header_font
    cell.fill  = header_fill
    cell.alignment = center_align

for row in precincts:
    ws1.append(list(row))

# Column widths
for col_idx, width in enumerate([12, 20, 12, 12, 10, 14, 12, 10, 10, 14], start=1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

# ──────────────────────────────────────────────────────────────────────────────
# Sheet 2: DA Race — Candidate Race
# ──────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("DA Race")

headers2 = [
    "Precinct_ID", "Precinct Name", "Registered", "Ballots Cast",
    "Mail JOHNSON", "Mail CHEN",  "Mail RAMIREZ",
    "In Person JOHNSON", "In Person CHEN", "In Person RAMIREZ",
    "JOHNSON", "CHEN", "RAMIREZ",
]

prec_da = [
    ("0001", "Oakdale North",    4200, 1717,  610, 490, 617,  310, 250, 310,   920, 740,  927),  # noqa
    ("0002", "Riverside Park",   3875, 1620,  520, 430, 670,  260, 220, 320,   780, 650,  990),
    ("0003", "Downtown Central", 5120, 2260,  850, 720, 690,  420, 360, 340,  1270,1080, 1030),
    ("0004", "Hillcrest East",   2900, 1230,  380, 280, 450,  195, 142, 233,   575, 422,  683),
    ("0005", "Westview Valley",  6300, 2700,  920, 780, 450,  460, 390, 230,  1380,1170,  680),
]

ws2.append(headers2)
for cell in ws2[1]:
    cell.font  = header_font
    cell.fill  = PatternFill("solid", fgColor="5D2A8A")
    cell.alignment = center_align

for row in prec_da:
    ws2.append(list(row))

for col_idx, width in enumerate([12,20,12,14,16,12,16,18,16,18,12,10,12], start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

wb.save(out_path)
print(f"Sample workbook written: {out_path}")
print(f"  Sheet 1: 'Measure A'  — {len(precincts)} precincts, ballot measure")
print(f"  Sheet 2: 'DA Race'    — {len(prec_da)} precincts, candidate race")
